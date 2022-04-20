from openpyxl import load_workbook


def insert_users(*args):
    """
    вносит нового юзера в БД

    :param args: принимает произволное колво аргументов
    """

    user_id = args[0]
    name = args[1]
    sex = args[2]
    grade = args[3]
    row = users_page.max_row + 1
    users_page.cell(row=row, column=1).value =user_id
    users_page.cell(row=row, column=2).value = name
    users_page.cell(row=row, column=3).value = sex
    users_page.cell(row=row, column=4).value = grade
    bd.save(database_filename)


def insert_sticker(keyword, sticker_id=None, reply_text=None):
    row = stickers_page.max_row + 1
    stickers_page.cell(row=row, column=1).value = keyword
    stickers_page.cell(row=row, column=2).value = sticker_id
    stickers_page.cell(row=row, column=3).value = reply_text
    bd.save(database_filename)
    stickers[keyword] = sticker_id
    replies[keyword] = reply_text


def in_database(user)-> bool:
    '''
    Return True if user_id is already in database
    '''
    for row in range(2, users_page.max_row + 1):
        if user == users_page.cell(row=row, column=1).value:
            return True
        return False




database_filename = 'DTbot.xlsx'
bd = load_workbook(database_filename)
stickers_page = bd['Stickers']
users_page = bd['Users']

users = {}
stickers = {}
replies = {}

for row in range(2, stickers_page.max_row + 1):
    keyword = stickers_page.cell(row=row, column=1).value
    sticker_id = stickers_page.cell(row=row, column=2).value
    reply_text = stickers_page.cell(row=row, column=3).value
    stickers[keyword] = sticker_id
    replies[keyword] = reply_text


if __name__ == '__main__':
    print(stickers)
    insert_sticker('до свидания', reply_text='и вам не хворать')